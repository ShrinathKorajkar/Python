import openpyxl, os
from openpyxl.cell.cell import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, LineChart, PieChart

os.chdir(os.path.dirname(__file__))
file = openpyxl.load_workbook('example.xlsx')
# print(file.get_sheet_names())
# sheet = file.get_sheet_by_name('Sheet1')
# print(dir(sheet))
print(type(file))
print(file.sheetnames)
sheet = file['Sheet1']
print(sheet)
print(sheet.title)
acSheet = file.active
print(acSheet)
print(acSheet['A1'])
print(acSheet['A1'].value)
spam = sheet['B1']
print(spam.row, spam.column, spam.value)
print(spam.coordinate)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# print(sheet.get_highest_row)
# print(column_index_from_string('A'))
print(sheet.max_row)
print(sheet.max_column)
print(get_column_letter(1))

print(tuple(sheet['A1':'C3']))
for row in sheet['A1':'C3']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('----END OF ROW----')

# for cell in sheet.columns[1]
for cell in sheet['B']:
    print(cell.value)

# WRITING IN EXCEL

print(file.sheetnames)
sheet2 = file['Sheet2']
sheet2.title = 'Phoenix'
file.create_sheet()
file.create_sheet(index=3, title='new sheet')
print(file.sheetnames)
# file.remove_sheet(file.get_sheet_by_name('Sheet1'))
file.remove(file['Sheet'])
print(file.sheetnames)
sheet2['A1'] = 'hello'
print(sheet2['A1'].value)
row = ('hello', 'world')
sheet2.append(row)

# FONT STYLE

# italic = Font(size=24, italic=True)
# style = cell_style(font=italic)
# sheet2['A1'].style = style
# print(sheet2['A1'].value)
sheet2['A1'].value = 'hello'
sheet2['B1'].value = 'world'
sheet2['A2'].value = 'hey'
sheet2['B2'].value = 'world'
sheet2['A1'].font = Font(size=24, italic=True)
sheet2['B1'].font = Font(name='Times New Roman', bold=True)

# FORMULAS AND DIMENSION
sheet2['A3'] = 100
sheet2['A4'] = 200
sheet2['A5'] = '=SUM(A3:A4)'
print(sheet2['A5'].value)

# wb = openpyxl.load_workbook('example_copy.xlsx', data_only=True)
# sheet2 = wb['Sheet2']
# print(sheet2['A5'].value)

sheet2['A6'].value = 'Tall row'
sheet2['B6'].value = 'wide row'
sheet2.row_dimensions[6].height = 70  # normal = 12.75 upto 409
sheet2.column_dimensions['B'].width = 70  # normal = 8.43 upto 255

# FREEZE PANES AND MERGING

sheet2.merge_cells('C1:D5')
sheet2['C1'].value = 'merged cells'
# sheet2.unmerge_cells('C1:D1')

sheet3 = file['Sheet3']
sheet3.freeze_panes = 'B2'

# CHARTS AND GRAPHS

for i in range(1, 11):
    sheet3['A' + str(i)] = i

ref = openpyxl.charts.Reference(sheet3, (1, 1), (10, 1))
series = openpyxl.charts.Series(ref, title='Series')
chart = openpyxl.charts.BarChart()
chart.append(series)
chart.drawing.top = 50
chart.drawing.left = 100
chart.drawing.width = 300
chart.drawing.height = 200
sheet3.add_chart(chart)

values = Reference(sheet3, min_col=1, min_row=1, max_col=1, max_row=10)

chart = BarChart()
chart.add_data(values)
chart.title = 'Series'
chart.x_axis.title = 'x axis'
chart.y_axis.title = 'y axis'
chart.height = 4
chart.width = 8
sheet3.add_chart(chart, 'C3')

chart = PieChart()
chart.add_data(values)
chart.title = 'Series'
chart.height = 4
chart.width = 8
sheet3.add_chart(chart, 'M3')

chart = LineChart()
chart.add_data(values)
chart.title = 'Series'
chart.x_axis.title = 'x axis'
chart.y_axis.title = 'y axis'
chart.height = 4
chart.width = 8
sheet3.add_chart(chart, 'H3')

file.save('example_copy.xlsx')